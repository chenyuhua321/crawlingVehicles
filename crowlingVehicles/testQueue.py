# -*- coding: utf-8 -*-
import threading
import Queue
import time
import openpyxl

class worker(threading.Thread):
    def __init__(self, queue):
        threading.Thread.__init__(self)
        self.queue = queue
        self.thread_stop = False

    def run(self):
        while not self.thread_stop:
            print("thread%d %s: waiting for tast" % (self.ident, self.name))
            try:
                task = q.get(block=True, timeout=2)  # 接收消息
            except Queue.Empty:
                print("Nothing to do! I will go home!")
                self.thread_stop = True
                break
            print("I am working")
            wb = openpyxl.load_workbook('queuetest.xlsx')
            ws = wb['173vehicle']
            ws.append(task[0])
            ws.append(task[1])
            wb.save('queuetest.xlsx')
            time.sleep(3)
            print("work finished!")
            q.task_done()                           # 完成一个任务
            res = q.qsize()                         # 判断消息队列大小(队列中还有几个任务)
            if res > 0:
                print("fuck! Still %d tasks to do" % (res))

    def stop(self):
        self.thread_stop = True

if __name__ == "__main__":
    q = Queue.Queue(3)
    wbo = openpyxl.Workbook()
    ws2 = wbo.create_sheet('173vehicle')
    wbo.save('queuetest.xlsx')# 创建队列（大小为3）
    worker = worker(q)                                    # 将队列加入类中
    worker.start()# 启动类
    alldata = []
    title = []
    entity = []
    title.append(1)
    title.append(2)

    entity.append('222')
    entity.append('333')
    alldata.append(title)
    alldata.append(entity)
    q.put(alldata)
    print("***************leader:wait for finish!")
    q.join()                                             # 等待所有任务完成
    print("***************leader:all task finished!")