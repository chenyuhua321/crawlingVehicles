# -*- coding: utf-8 -*-
import time
import re
import json
from selenium import webdriver
from bs4 import BeautifulSoup
from realGetVin import getVin
import threading
import random
from threading import Thread
import codecs
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from concurrent.futures import ThreadPoolExecutor, as_completed,ProcessPoolExecutor
import openpyxl
import Queue
import redis
from NeCodeTest import findByCode

import sys
reload(sys)
sys.setdefaultencoding('utf8')
'''爬取工信部车辆信息'''
class findVin(threading.Thread):
    def __init__(self,queue, time2wait=10):
        '''self.chromeOptions = webdriver.ChromeOptions()
        self.chromeOptions.add_argument('--headless')
        self.chromeOptions.add_argument('--disable-gpu')
        self.driver = webdriver.Chrome(chrome_options=self.chromeOptions)'''
        '''self.firefoxOptions = webdriver.FirefoxOptions()
        # 设置代理
        #self.firefoxOptions.add_argument('--headless')
        #self.firefoxOptions.add_argument('--disable-gpu')
        # 一定要注意，=两边不能有空格，不能是这样--proxy-server = http://202.20.16.82:10152
        self.driver = webdriver.Firefox(firefox_options = self.firefoxOptions)

        self.driver.get('http://www.miit-eidc.org.cn/miitxxgk/gonggao/xxgk/queryByPc?pc=173&querylb=cp&qyinfolb=')
        self.driver.implicitly_wait(5)
        self.wait = WebDriverWait(self.driver, time2wait)
        self.driver.find_element_by_id("query_qymc_input").send_keys(u"公司")
        self.driver.find_element_by_id("query_button").click()'''

        self.count =1
        self.vehicount =1
        threading.Thread.__init__(self)
        self.queue = queue
        self.thread_stop = False
        self.sleepCount = 0
        self.defeatCount =0
        self.defeatFlag = True


    def printclick(self,allTypes):
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        query_result = soup.find('div',{'id':'query_result'})
        d = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, 'query_page_count')))
        pagesize = int(soup.find('input', {'id':'query_page_count'})['value'])
        pagenow = '0'
        while int(pagenow) != self.count:
            try:
                e = WebDriverWait(self.driver, 10,0.5).until(EC.presence_of_element_located((By.CLASS_NAME, 'query_result_now_page')))
                pagenow = e.text
            except Exception:
                time.sleep(1)
                e = WebDriverWait(self.driver, 10,0.5).until(EC.presence_of_element_located((By.CLASS_NAME, 'query_result_now_page')))
                pagenow = e.text
        tbody = query_result.find('tbody')
        trs = tbody.find_all('tr')
        alltypeFile = open("173allayne", "a+")
        Num = []
        Id = []
        batch = []
        for tr in trs:
            tds = tr.find_all('td')
            td = tds[3]
            a = td.find('a')
            if a != None:
                hrefs = re.findall('"([^"]+)"', a['href'])
                Num.append(hrefs[0])
                Id.append(hrefs[1])
                batch.append(hrefs[2])
                alltypeFile.write(hrefs[0]+" "+hrefs[1]+" "+hrefs[2]+"\n")
        alltypeFile.close()
        allType = []
        allType.append(Num)
        allType.append(Id)
        allType.append(batch)
        allTypes.append(allType)
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'query_result_cph_link')))
        query_result_cph_links = self.driver.find_elements_by_class_name('query_result_cph_link')
        find_links = self.driver.find_elements_by_link_text('查看')
        rightlinks = []
        for query_result in query_result_cph_links:
            if query_result not in find_links:
                rightlinks.append(query_result)

       # for query_result in rightlinks:
           # query_result.click()
       # handles =  self.driver.window_handles
        #self.clicktab(handles)
        #executor = ThreadPoolExecutor(max_workers=5)
        #all_task = [ executor.submit(self.clicktab,query_result) for query_result in query_result_cph_links]


        if(pagesize != self.count):
            input = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, 'query_goto_page')))
            #input = driver.find_element_by_id('query_goto_page')
            input.clear()
            self.count = self.count +1
            input.send_keys(self.count)
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.LINK_TEXT, 'GO'))).click()
            #driver.find_element_by_link_text('GO').click()
            self.printclick(allTypes)
        else:
            print len(allTypes)
        return allTypes
    def realfindVin(self,tag,gid,batch):
        pool = redis.ConnectionPool(host='localhost', port=6379)
        red = redis.Redis(connection_pool=pool)
        vehicledata = []
        time_begin = time.time()
        while red.scard('NECaptchaValidate') == 0:
            print '等一个有缘人'
            time.sleep(2)
        if red.scard('NECaptchaValidate') > 0:
            NECaptchaValidate = red.spop('NECaptchaValidate')
            print NECaptchaValidate
            vehicledata = findByCode(tag,gid,batch,NECaptchaValidate)
        print vehicledata
        print '--------------------------'
        time_end = time.time()
        alltime = time_end - time_begin
        print ('time:', alltime)
        title = []
        entity = []
        for vehi in vehicledata[1]:
            title.append(vehi.decode('utf-8'))
        for realdata in vehicledata[2]:
            entity.append(realdata.decode('utf-8'))
        print str(self.vehicount)+'成功'
        self.vehicount = self.vehicount+1
        alldata = []
        alldata.append(title)
        alldata.append(entity)
        besize = q.qsize()
        aftersize = besize
        while besize == aftersize:
            q.put(alldata,block=True, timeout=None)
            aftersize = q.qsize()
        print ('==================')
    def finVin(self,tag,gid,batch):
        print 'start'+ gid

        time_begin = time.time()
        vehicledata = None
        time.sleep(2)
        if self.defeatFlag:
            while vehicledata == None :
                vehicledata = getVin(tag,gid,batch)
            if vehicledata != None:
                while len(vehicledata)<3:
                    vehicledata = getVin(tag,gid,batch)
                    self.defeatCount = self.defeatCount +1
                    print str(self.defeatCount)+'失败次数'
                    if self.defeatCount > 20:
                        self.defeatFlag=False
                        time.sleep(300)
                        self.defeatCount = 0
        print vehicledata
        print '--------------------------'
        time_end = time.time()
        alltime = time_end - time_begin
        print ('time:', alltime)
        title = []
        entity = []
        for vehi in vehicledata[1]:
            title.append(vehi.decode('utf-8'))
        for realdata in vehicledata[2]:
            entity.append(realdata.decode('utf-8'))
        print str(self.vehicount)+'成功'
        self.vehicount = self.vehicount+1
        alldata = []
        alldata.append(title)
        alldata.append(entity)
        if q.qsize()>5:
            if self.defeatFlag:
                self.defeatFlag =False
                aa = threading.Thread(target=self.run())
                self.defeatFlag =True

        self.thread_stop =False
        besize = q.qsize()
        aftersize = besize
        while besize == aftersize:
            q.put(alldata,block=True, timeout=None)
            aftersize = q.qsize()
        print ('==================')


    def toexcel(self,data):
        try:
            '''wbo = openpyxl.Workbook()
            ws2 = wbo.create_sheet('173vehicle')
            wbo.save('173allvehicle.xlsx')'''
            tag = []
            gid =[]
            pc = []
            '''for da in data:
                for ln in range(len(da[0])):
                    tag.append(da[0][ln])
                    gid.append(da[1][ln])
                    pc.append(da[2][ln])'''
            tag = data[0]
            gid = data[1]
            pc=data[2]
            print len(tag)
            executor = ThreadPoolExecutor(max_workers=5)
            all_task = [ executor.submit(self.realfindVin,tag[lent],gid[lent],pc[lent]) for lent in range(len(tag))]

            self.start()
        except Exception as e:
            print e
    def openalltype(self):
        f = open('173allayne','rb')
        data = []
        tag = []
        gid =[]
        pc = []
        while True:
            lines = f.readline()
            print lines
            tag.append(lines[0:1])
            gid.append(lines[2:9])
            pc.append(lines[10:13])
            if not lines:
                break
        f.close()
        data.append(tag)
        data.append(gid)
        data.append(pc)
        return data

    def run(self):
        while True:
            print("thread%d %s: waiting for tast" % (self.ident, self.name))
            try:
                task = q.get(block=True, timeout=2)  # 接收消息
            except Queue.Empty:
                print("Nothing to do! I will go home!")
                continue
            print("I am working")
            wb = openpyxl.load_workbook('173allvehicle.xlsx')
            ws = wb['173vehicle']
            ws.append(task[0])
            ws.append(task[1])
            wb.save('173allvehicle.xlsx')
            time.sleep(3)
            print("work finished!")
            q.task_done()                           # 完成一个任务
            res = q.qsize()                         # 判断消息队列大小(队列中还有几个任务)
            if res > 0:
                print("Still %d tasks to do" % (res))

    def stop(self):
        self.thread_stop = True



if __name__ == '__main__':
    print 'begin find'
    q = Queue.Queue()
    a = findVin(q)
    alltype = []
    time.sleep(1)
    allbegin = time.time()
    data = a.openalltype()
    #data = a.printclick(alltype)
    a.toexcel(data)                                  # 将队列加入类中
    allend = time.time()
    alltime = allend - allbegin

    print ('alltime:',alltime)



