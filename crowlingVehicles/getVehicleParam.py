# -*- coding: utf-8 -*-
import time
import re
import json
from selenium import webdriver
from bs4 import BeautifulSoup
from realGetVin import getVin
import threading
import requests
from fake_useragent  import UserAgent
import random
from threading import Thread
import codecs
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from concurrent.futures import ThreadPoolExecutor, as_completed,ProcessPoolExecutor
import openpyxl
import Queue
import redis
from NeCodeTest import findByCode

import sys
reload(sys)
sys.setdefaultencoding('utf8')
'''爬取工信部车辆信息'''
class getVehicleParam(threading.Thread):
    def __init__(self,queue,batch, time2wait=10):
        threading.Thread.__init__(self)
        print batch
        self.batch = batch
        self.queue = queue
        self.thread_stop = False
        self.sleepCount = 0
        self.defeatCount =0
        self.defeatFlag = True

    def getParam(self,param):
        pool = redis.ConnectionPool(host='localhost', port=6379)
        red = redis.Redis(connection_pool=pool)
        param = red.spop(self.batch+'vehicle')
        url = 'http://www.chinacar.com.cn/Home/GonggaoSearch/GonggaoSearch/search_param/id/'+param
        ua = UserAgent()
        useragent = ua.random
        headers = {
            'Accept': 'text/html, */*; q=0.01',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'Connection': 'keep-alive',
            'Cookie': 'PHPSESSID=511bpb2a85rho27t3flpehqlr1; Hm_lvt_6c1a81e7deb77ce536977738372f872a=1568626376,1569489687; rel_search=1; relv=1; vin_cookie=0506; news_view_ids=%7C88215%7C; BDPCIEXP=30; BDTUJIAID=f034d8969b597e7ffc0b29cd54e18993; clcp_list=796130%7C437948%7C428562%7C428680%7C796132%7C795730%7C795736%7C; ck_gg_1=y; Hm_lpvt_6c1a81e7deb77ce536977738372f872a=1569585866',
            'Host': 'www.chinacar.com.cn',
            'Referer': 'http://www.chinacar.com.cn/ggcx_new/search_view.html?id='+param,
            'User-Agent':useragent,
            'X-Requested-With': 'XMLHttpRequest'
        }
        statue = 1
        while not statue == 200:
            response = requests.get(url,headers=headers)
            statue = response.status_code
        soup = BeautifulSoup(response.text, 'html.parser')
        trs = soup.find_all('tr')
        data = []
        for tr in trs :
            tds = tr.find_all('td')
            for td in tds :
                data.append(td.get_text())
        formatDatas = []
        for lenth in range(len(data)):
            if not lenth in [0,33,38,47,114,123,132]:
                realdata = str(data[lenth])
                formatDatas.append(realdata.decode(encoding='utf-8'))
        firstDatas = formatDatas[0:110]
        firstTitle =firstDatas[::2]
        firstEntity = firstDatas[1::2]
        for lenth in range(110,114):
            firstTitle.append(formatDatas[lenth])
            firstEntity.append(formatDatas[lenth+4])
        endDatas = formatDatas[118:]
        endTitle = endDatas[::2]
        endEntity = endDatas[1::2]
        title = firstTitle+endTitle
        entity = firstEntity+endEntity
        alldata = []
        alldata.append(title)
        alldata.append(entity)
        besize = q.qsize()
        q.put(alldata,block=True, timeout=None)
        q.join()
    def toexcel(self):
        try:
            pool = redis.ConnectionPool(host='localhost', port=6379)
            red = redis.Redis(connection_pool=pool)
            while red.scard(self.batch+'vehicle') == 0:
                print '等一个有缘人'
                time.sleep(2)
            data = red.smembers(self.batch+'vehicle')

            executor = ThreadPoolExecutor(max_workers=100)
            all_task = [ executor.submit(self.getParam,d) for d in data]
        except Exception as e:
            print e

    def run(self):
        while not self.thread_stop:
            tasks = []
            while not q.empty():
                task = q.get(block=True, timeout=2)# 接收消息
                q.task_done()
                tasks.append(task)
            if len(tasks) >0:
                print("I am working")
                print str(self.batch)+'allvehicle.xlsx'
                try:
                    wb = openpyxl.load_workbook('vehicle/'+str(self.batch)+'allvehicle.xlsx')
                    ws = wb[str(self.batch)+'vehicle']
                    for task in tasks:
                        title =[]
                        for ta in task[1]:
                            title.append(ta.decode(encoding='utf-8'))
                        ws.append(title)
                    wb.save('vehicle/'+str(self.batch)+'allvehicle.xlsx')
                except IOError:
                    wb = openpyxl.Workbook()
                    ws = wb.create_sheet(str(self.batch)+'vehicle')
                    for lenth in range(len(tasks)):
                        title =[]
                        for ta in tasks[lenth][0]:
                            title.append(ta.decode(encoding='utf-8'))
                        entity =[]
                        for en in tasks[lenth][1]:
                            entity.append(en.decode(encoding='utf-8'))
                        if lenth == 0:
                            ws.append(title)
                            ws.append(entity)
                        else:
                            ws.append(entity)
                    wb.save('vehicle/'+str(self.batch)+'allvehicle.xlsx')

                time.sleep(3)
                print("work finished!")
                                   # 完成一个任务
            print("Nothing to do! I will go home!")
            continue
    def stop(self):
        self.thread_stop = True



if __name__ == '__main__':
    print 'begin find'
    data = []
    q = Queue.Queue()
    allbegin = time.time()
    g = getVehicleParam(q,'219')
    g.toexcel()
    g.start()
    time.sleep(1)                                  # 将队列加入类中
    allend = time.time()
    alltime = allend - allbegin

    print ('alltime:',alltime)